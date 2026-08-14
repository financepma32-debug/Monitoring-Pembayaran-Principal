"""
VALIDASI COA FBL3N vs MASTER COA — v6 (LOKAL, tanpa Supabase/website)
=======================================================
Cara pakai:
  1. Double-click RUN.bat (atau: python validasi_coa_fbl3n.py)
  2. Output tersimpan di folder OUTPUT\ (di sebelah script ini)
  3. Salinan output JUGA otomatis disimpan ke folder kerja divisi di
     drive D: (lihat SIMPAN_KE_FOLDER_DIVISI di bawah) -- menggantikan
     sinkronisasi ke Supabase/website yang dipakai versi sebelumnya.

Tidak ada satu baris pun logika evaluasi COA (evaluate(), validate_dht(),
saran(), proses_sheet(), dll) yang diubah dari versi sebelumnya -- yang
dihapus di versi ini murni bagian login + kirim data ke Supabase (yang
sebelumnya opsional lewat dashboard_biaya_config.ini). Skrip ini portable:
bisa disalin/dijalankan dari folder mana pun, di komputer mana pun yang
sudah di-SETUP.bat -- outputnya tetap selalu tersimpan di OUTPUT\ lokal
DAN di folder divisi pada drive D: seperti diatur di bawah.
"""

import openpyxl, re, os, io, shutil
from difflib import SequenceMatcher
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles.colors import Color as OXColor
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from datetime import datetime

# ==============================
#  PATH
# ==============================
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR  = os.path.join(BASE_DIR, "OUTPUT")
INPUT_FBL3N = os.path.join(BASE_DIR, "FBL3N.xlsx")
INPUT_MASTER= os.path.join(BASE_DIR, "MASTER_COA.xlsx")
_TS         = datetime.now().strftime("%Y%m%d_%H%M%S")
_user_candidates = [
    os.path.join(BASE_DIR, "master_user.xlsx"),
    os.path.join(BASE_DIR, "master user.xlsx"),
    os.path.join(os.path.dirname(BASE_DIR), "master_user.xlsx"),
    os.path.join(os.path.dirname(BASE_DIR), "master user.xlsx"),
    r"D:\PROJECT FAD\KK ACC\master_user.xlsx",
    r"D:\PROJECT FAD\KK ACC\master user.xlsx",
    r"D:\PROJECT FAD\master_user.xlsx",
    r"D:\PROJECT FAD\master user.xlsx",
]
INPUT_USER = next((p for p in _user_candidates if os.path.exists(p)),
                  os.path.join(BASE_DIR, "master_user.xlsx"))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, f"FBL3N_Validated_{_TS}.xlsx")

# ==============================
#  TUJUAN SALINAN OUTPUT DI DRIVE D:
#  Menggantikan sinkronisasi ke Supabase/website versi sebelumnya --
#  output tetap tersimpan seperti biasa di OUTPUT\ lokal, DAN sebuah
#  salinannya otomatis ditaruh di:
#    ROOT_PROJECT_FAD\DIVISI_TUJUAN\FOLDER_TUJUAN\OUTPUT\<waktu_proses>\
#  Tiap run dapat folder sendiri (dinamai sesuai waktu proses) supaya
#  tidak menumpuk jadi satu tumpukan file di folder kerja divisi -- tim
#  lain (mis. P&L) tinggal buka folder ini langsung, tanpa login web.
#  Ganti 2 baris DIVISI_TUJUAN/FOLDER_TUJUAN ini kalau salinan script
#  dipakai di folder kerja/divisi yang lain.
# ==============================
ROOT_PROJECT_FAD = r"D:\PROJECT FAD"
DIVISI_TUJUAN    = "ACCOUNTING"
FOLDER_TUJUAN    = "1. KK REVIEW JURNAL"

# ==============================
#  STYLE
# ==============================
FILL_HDR   = PatternFill("solid", fgColor="C0C0C0")
FILL_NONE  = PatternFill(fill_type=None)
FILL_OK    = PatternFill("solid", fgColor="D9EAD3")
FILL_NO    = PatternFill("solid", fgColor="F4CCCC")
FILL_WARN  = PatternFill("solid", fgColor="FFF2CC")
FILL_AKSES = PatternFill("solid", fgColor="FCE5CD")
FILL_DHT   = PatternFill("solid", fgColor="CFE2F3")
FILL_FMT   = PatternFill("solid", fgColor="FFE599")
FILL_CC_NO = PatternFill("solid", fgColor="F4CCCC")
FILL_CC_OK = PatternFill("solid", fgColor="D9EAD3")
FILL_FMT_NO= PatternFill("solid", fgColor="FCE5CD")
ALIGN_TOP  = Alignment(vertical="top")
BORDER_HDR = Border(left=Side(border_style="thin"))
BORDER_NON = Border()
ROW_H1     = 12.75
ROW_H2     = 14.25

def fnt(sz=10):
    return Font(name="Aptos Narrow", size=sz, bold=False, color=OXColor(theme=1))

# ==============================
#  AUTHORIZED
# ==============================
def authorized(username: str) -> bool:
    """Hanya P000_xxx (HO/Finance) dan SAP_SYSTEM yang boleh akses GL 51 & 6."""
    u = re.sub(r'[\s.]', '', str(username)).upper()
    if 'SAP_SYSTEM' in u or u.startswith('SAP'):
        return True
    m = re.match(r'^P(\d+)', u)
    if m:
        return int(m.group(1)) == 0
    return False

# ==============================
#  TOKENIZER
# ==============================
_SW = {'DAN','ATAU','UNTUK','YANG','DARI','KE','DI','PER','DGN','YG','UTK','DLL',
       'HO','SLS','NSLS','DPS','MM','YY','DD','NO','NOMOR','TSB','SBB'}

def tok(text: str) -> set:
    raw = re.split(r'[\s_\-/(),\.;:&+\[\]<>]+', text.upper())
    return {t for t in raw if len(t) >= 2 and t not in _SW and not t.isdigit()}

# ==============================
#  NOPOL DETECTOR
# ==============================
def has_nopol(text: str) -> bool:
    n = re.sub(r'[-_/.]', ' ', text.upper())
    n = re.sub(r'([A-Z]{1,2}\d{1,5}[A-Z]{1,3})(KM)', r'\1 \2', n)
    return (
        bool(re.search(r'\b[A-Z]{1,2}\d{1,5}[A-Z]{1,3}\b', n)) or
        bool(re.search(r'\b[A-Z]{1,2}\s\d{1,4}\s[A-Z]{1,3}\b', n)) or
        bool(re.search(r'\b[A-Z]{1,2}\d{4,5}(?=[\s,])', n))
    )

# ==============================
#  FUZZY MATCH
# ==============================
def _sim(a, b):
    return SequenceMatcher(None, a, b).ratio() * 100

def _psim(a, b):
    if len(a) > len(b): a, b = b, a
    best = 0.0
    for i in range(len(b) - len(a) + 1):
        s = SequenceMatcher(None, a, b[i:i+len(a)]).ratio() * 100
        if s > best: best = s
    return best

def fuzzy_hit(tokens: set, pool: list, threshold=72) -> tuple:
    for t in tokens:
        if len(t) < 2: continue
        best_kw, best_sc = None, 0.0
        for kw in pool:
            s = _sim(t, kw)
            if s > best_sc: best_sc, best_kw = s, kw
            if len(t) <= 5:
                ps = _psim(t, kw)
                if ps > best_sc: best_sc, best_kw = ps, kw
        if best_sc >= threshold and best_kw:
            return True, f"{t}~{best_kw}({best_sc:.0f}%)"
    return False, ""

# ==============================
#  DHT KW MAP
# ==============================
DHT_KW_MAP = {
    'BBM KENDARAAN NIAGA'   : {'52010101'},
    'BBM KEND NON NIAGA'    : {'54010101'},
    'PARKIR&TOL KEND'       : {'52010102','52010103'},
    'PAR&TOL KEND NON'      : {'54010102'},
    'RETRIBUSI KEND NIAGA'  : {'52010103'},
    'SEWA KENDARAAN NIAGA'  : {'52010301'},
    'SEWA KNDRN NIAGA LAIN' : {'52010309'},
    'SEWA KEND NIAGA LAIN'  : {'52010309'},
    'UPAH HARIAN LEPAS'     : {'52010401'},
    'POCOAN HELPER'         : {'52010405','52010401'},
    'ONGKOS KULI'           : {'52010402'},
    'ANGKUT INTRAG'         : {'52010403'},
    'OWH LAINNYA'           : {'52010406'},
    'PEM KEND NIAGA'        : {'52010201'},
    'BEBAN PML KEND NIAGA'  : {'52010201'},
    'ATK & CETAKAN'         : {'55010101'},
    'POS DAN LAYANAN DOK'   : {'55010103'},
    'RUMAH TANGGA KANTOR'   : {'55010201'},
    'BY LISTRIK'            : {'55020001'},
    'ADM BANK'              : {'72010001'},
    'ADMINISTRASI BANK'     : {'72010001'},
    'BY HOTEL'              : {'55080201'},
    'LEBIH BAYAR'           : {'71060002'},
    'KURANG BAYAR'          : {'71060001'},
    'ASS KENDARAAN MOBIL'   : {'52020001'},
}

# ==============================
#  VALIDATE DHT
#  [FIX B] DHT MISMATCH -> TIDAK SESUAI
# ==============================
def validate_dht(gl: str, dht: str, coa_map: dict) -> str:
    """
    Return 'OK' | 'MISMATCH' | 'NOINFO'
    Cek DHT terhadap NAMA GL SAP dari master COA (kolom 4).
    """
    if not dht or not gl:
        return 'NOINFO'
    dht_u = str(dht).strip().upper()
    if not dht_u or len(dht_u) < 3 or dht_u.isdigit():
        return 'NOINFO'

    # Cek apakah DHT sesuai dengan COA ini
    expected_dht = coa_map.get(gl, {}).get('dht', '')
    if expected_dht and expected_dht in dht_u:
        return 'OK'

    # Cek apakah DHT lebih cocok ke COA lain
    for other_gl, data in coa_map.items():
        if other_gl == gl:
            continue
        other_dht = data.get('dht', '')
        if other_dht and len(other_dht) >= 4 and other_dht in dht_u:
            return 'MISMATCH'

    return 'NOINFO'

# ==============================
#  LOAD MASTER COA (Dinamis dari file)
#  [FIX C] Auto-extract keywords dari sub-items
#  Update master COA -> otomatis terbaca di run berikutnya
# ==============================
def _kw_add(kw_set: set, text):
    if not text: return
    for w in re.sub(r'[^A-Z0-9\s]', ' ', str(text).upper()).split():
        if len(w) >= 3 and not w.isdigit():
            kw_set.add(w)

def _kw_pattern(kw_set: set, text):
    if not text: return
    for p in re.split(r'[_/,;]', str(text).upper()):
        first = p.strip().split()[0] if p.strip().split() else ''
        if len(first) >= 3 and not first.isdigit():
            kw_set.add(first)

def load_master(path: str) -> dict:
    """
    Baca MASTER_COA.xlsx dan extract keywords per COA secara dinamis.
    Setiap run membaca ulang file — tambah sub-item di master langsung terbaca.
    Kolom: GL(1) | Nama GL(2) | GL SAP(3) | NAMA GL SAP(4) | no_sub(5) |
           Description/sub-item(6) | ... | Std Penulisan(11) | Contoh(12)
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    coa_map    = {}
    current_gl = None

    for row in ws.iter_rows(min_row=8, values_only=True):
        gl_sap   = row[2]   # GL SAP (index 2)
        nama_sap = row[3]   # NAMA GL SAP = DHT expected (index 3)
        desc_gl  = row[5]   # Description GL / sub-item text (index 5)
        sub_id   = row[4]   # nomor sub-item (index 4)
        std      = row[10]  # Standard Penulisan (index 10)
        contoh   = row[11]  # Contoh (index 11)

        gl_str = str(gl_sap).strip() if gl_sap else ''
        if re.match(r'^[5-7]\d{7}$', gl_str):
            current_gl = gl_str
            kws = set()
            _kw_add(kws, nama_sap)
            _kw_add(kws, desc_gl)
            _kw_pattern(kws, std)
            _kw_pattern(kws, contoh)
            coa_map[current_gl] = {
                'nama'     : str(nama_sap).strip() if nama_sap else '',
                'dht'      : str(nama_sap).strip().upper() if nama_sap else '',
                'desc'     : str(desc_gl).strip() if desc_gl else '',
                'std'      : str(std).strip() if std else '',
                'keywords' : kws,
                'tokens'   : kws,
                'sub_items': [],
            }
            continue

        if current_gl and (sub_id is not None or desc_gl):
            if desc_gl:
                coa_map[current_gl]['sub_items'].append(str(desc_gl).strip())
                _kw_add(coa_map[current_gl]['keywords'], desc_gl)
            _kw_pattern(coa_map[current_gl]['keywords'], std)
            _kw_pattern(coa_map[current_gl]['keywords'], contoh)

    wb.close()
    return coa_map

def coa_label(gl: str, coa_map: dict) -> str:
    """Return 'GL - Nama COA'."""
    nama = coa_map.get(gl, {}).get('nama', '')
    return f'{gl} - {nama}' if nama else gl

# ==============================
#  LOAD MASTER USER (Dinamis dari file)
#  Update master user -> otomatis terbaca di run berikutnya
# ==============================
def load_master_user(path: str) -> dict:
    """Baca master_user.xlsx. Sebagian baris tidak punya User Name (kolom A
    kosong) -- ini daftar Cost Center HO/kantor pusat (mis. 'P01H060002 -
    HO-ACCOUNTING') yang tidak dipegang satu orang PIC tertentu, tapi TETAP
    perlu masuk supaya build_cc_index() bisa mengenalinya (lihat cek_cc()
    utk baris HO/Finance P000_xxx). Baris begini disimpan pakai kunci
    sintetis '__CC__{cost center}' -- tidak akan pernah cocok dgn username
    SAP asli (_lookup_user selalu mencari berdasar pola P<angka>...), jadi
    aman tidak akan salah tertaut ke user manapun."""
    if not os.path.exists(path):
        return {}
    wb  = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws  = wb.active
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        uname, kode, cc, nama, level = row[0], row[1], row[2], row[3], row[4]
        status  = row[5] if len(row) > 5 else ''
        channel = row[6] if len(row) > 6 else ''
        if not uname and not cc:
            continue
        key = str(uname).strip().upper() if uname else f'__CC__{str(cc).strip().upper()}'
        out[key] = {
            'cc'     : str(cc).strip()      if cc      else '',
            'kode'   : str(kode).strip()    if kode    else '',
            'nama'   : str(nama).strip()    if nama    else '',
            'level'  : level or 0,
            'status' : str(status).strip()  if status  else '',
            'channel': str(channel).strip() if channel else '',
        }
    wb.close()
    return out

def build_cc_index(user_map: dict) -> dict:
    """cc -> {'kode':..., 'nama':...}, dibangun dari SEMUA baris master_user.xlsx
    (bukan per-username) -- dipakai utk validasi baris HO/Finance (P000_xxx):
    mereka tidak terdaftar sendiri di master (lihat _lookup_user), jadi tidak
    bisa dicek lewat "CC milik user ini", tapi Cost Center yang mereka pakai
    seharusnya tetap salah satu Cost Center depo yang SUDAH terdaftar --
    kalau tidak ketemu di sini, berarti Cost Center baru yang belum
    dimasukkan ke master."""
    idx = {}
    for info in user_map.values():
        cc = (info.get('cc') or '').strip().upper()
        if cc and cc not in idx:
            idx[cc] = {'kode': info.get('kode', ''), 'nama': info.get('nama', ''),
                       'status': info.get('status', ''), 'channel': info.get('channel', '')}
    return idx


def _lookup_user(username: str, user_map: dict):
    """Handle semua format username FBL3N: P191_EDS, P238.239_KSR, P391_KSR, dll."""
    if not username: return None
    u = str(username).strip().upper()
    if u in user_map: return user_map[u]
    m = re.match(r'^P(\d+)', u)
    if not m: return None
    kode_num = int(m.group(1))
    kode_str = f'P{kode_num}'
    if kode_num == 0: return None
    for suf in ('_EDS', '_AOS_EDS', '_KSR_MT', '_AOS_MT', '_SA_MT',
                '_KSR', '_SPV', '_SA', '_LOG', '_LOG2'):
        if kode_str + suf in user_map:
            return user_map[kode_str + suf]
    return None

def cek_cc(username: str, cc_fbl3n: str, user_map: dict, cc_index: dict) -> tuple:
    """
    Return (keterangan_kolom_D, fill, channel, paksa_status)
    - Sesuai                  -> CC user (non-HO) benar
    - 'HO/Finance'            -> P000, Cost Center dikenali (ada di master) -> Sesuai
    - 'Profit Center Baru'    -> Cost Center TIDAK ada di master mana pun, ATAU user
                                 non-HO ini sendiri tidak terdaftar (Tidak Sesuai)
    - 'Profit Center Non Aktif (NAMA)' -> Cost Center ADA di master tapi kolom Status-nya
                                 'Non Aktif' -> paksa_status='USER TIDAK SAH'
    - 'P01Gxxxxxx (NAMA)'     -> CC yang seharusnya utk user non-HO ini (Tidak Sesuai)
    - kosong                  -> SAP_SYSTEM, skip
    paksa_status: None (biarkan status hasil evaluate() apa adanya) atau
                  'USER TIDAK SAH' (paksa status akhir baris ini, dipakai
                  proses_sheet utk override kolom STATUS).
    """
    if not username: return ('', None, '', None)
    u = str(username).strip().upper()
    if u.startswith('SAP_') or u == 'SAP_SYSTEM': return ('', None, '', None)
    m_p = re.match(r'^P(\d+)', u)
    if m_p and int(m_p.group(1)) == 0:
        # HO/Finance boleh posting ke Cost Center depo mana pun -- tidak bisa
        # dicek lewat "CC milik user ini" (P000 tidak terdaftar per-orang di
        # master_user.xlsx), jadi dicek lewat cc_index: apakah Cost Center
        # yang dipakai memang salah satu yang SUDAH terdaftar di master.
        cc_key = str(cc_fbl3n).strip().upper() if cc_fbl3n else ''
        if not cc_key or cc_key in ('NONE', 'NAN'):
            return ('HO/Finance', FILL_CC_OK, '', None)
        entry = cc_index.get(cc_key)
        if not entry:
            return ('Profit Center Baru', FILL_CC_NO, '', None)
        channel = entry.get('channel', '')
        if (entry.get('status') or '').strip().upper() == 'NON AKTIF':
            return (f'Profit Center Non Aktif ({entry.get("nama","")})',
                     FILL_AKSES, channel, 'USER TIDAK SAH')
        return ('HO/Finance', FILL_CC_OK, channel, None)
    info = _lookup_user(username, user_map)
    if not info: return ('Profit Center Baru', FILL_CC_NO, '', None)
    cc_master = info.get('cc', '')
    nama_depo = info.get('nama', '')
    channel   = info.get('channel', '')
    if (info.get('status') or '').strip().upper() == 'NON AKTIF':
        return (f'Profit Center Non Aktif ({nama_depo})',
                 FILL_AKSES, channel, 'USER TIDAK SAH')
    if not cc_fbl3n or cc_fbl3n in ('None', '', 'nan'):
        return (f'{cc_master} ({nama_depo})', FILL_CC_NO, channel, None)
    if cc_fbl3n.strip().upper() == cc_master.strip().upper():
        return ('Sesuai', FILL_CC_OK, channel, None)
    else:
        return (f'{cc_master} ({nama_depo})', FILL_CC_NO, channel, None)

# ==============================
#  SARAN COA
# ==============================
def saran(gl: str, tu: str, doc: str, coa_map: dict) -> str:
    if not tu: return ""
    if gl.startswith('51') or gl.startswith('6'): return ""
    doc0  = str(doc)[0] if doc else ''
    tu_up = tu.upper()

    if gl == '52010309' and doc0 == '4':
        return coa_label('52010301', coa_map)
    if gl == '52010301' and doc0 != '4':
        return coa_label('52010309', coa_map)

    # GRAB/GOJEK ke bank -> 55010201 (transport kasir)
    if any(k in tu_up for k in {'GRAB','GOJEK','OJEK'}) and \
       any(k in tu_up for k in {'BANK','SETOR','BRILINK','BRI','MANDIRI','BCA','BSI'}):
        return coa_label('55010201', coa_map)

    # GRAB/GOJEK tanpa bank -> 55080301
    if any(k in tu_up for k in {'GRAB','GOJEK','OJEK','TAXI','TAKSI'}):
        return coa_label('55080301', coa_map)

    # BBM + MOTOR -> 54010101
    if 'BBM' in tu_up and any(k in tu_up for k in {'MOTOR','SEWA MOTOR','SW MTR'}):
        return coa_label('54010101', coa_map)

    # [FIX E] KEAMANAN KANTOR/SATPAM/SECURITY di GL lain -> saran 55050009
    if any(k in tu_up for k in {'KEAMANAN KANTOR','SATPAM','SECURITY','SECURITI',
                                  'JAGA MALAM','PETUGAS KEAMANAN'}):
        return coa_label('55050009', coa_map)

    # Cari COA terbaik dari keyword match
    best_gl, best_score = None, 0
    tok_set = set(re.sub(r'[^A-Z0-9\s]', ' ', tu_up).split())
    for cgl, data in coa_map.items():
        if cgl == gl: continue
        kws = data.get('keywords', set())
        score = len(tok_set & kws)
        if score > best_score:
            best_score, best_gl = score, cgl
    if best_gl and best_score >= 2:
        return coa_label(best_gl, coa_map)

    return ""

# ==============================
#  ENGINE EVALUASI
# ==============================
def evaluate(gl, text_p, doc, username, amount, coa_map, dht=''):
    """
    Return: (status, remarks, saran_coa)
    status  : SESUAI | TIDAK SESUAI | USER TIDAK SAH
    remarks : Teks Kosong | COA-Text-Header Tidak Relevan |
              Koreksi COA | Cek Standar Penulisan | User Tidak Sah | ''
    """
    tu   = str(text_p).strip().upper() if text_p else ''
    doc  = str(doc).strip() if doc else ''
    doc0 = doc[0] if doc else ''
    dhtu = str(dht).strip().upper() if dht else ''
    amt  = float(amount) if isinstance(amount, (int, float)) else 0.0
    absa = abs(amt)
    t    = tok(tu)

    # [FIX B] ok() -> DHT MISMATCH langsung TIDAK SESUAI
    def ok():
        """Text SESUAI COA. Tapi jika DHT menunjuk GL lain -> TIDAK SESUAI."""
        dht_st = validate_dht(gl, dht, coa_map)
        if dht_st == 'MISMATCH':
            return ('TIDAK SESUAI', 'COA-Text-Header Tidak Relevan', '')
        return ('SESUAI', '', '')

    # 1. TEXT KOSONG
    if not tu or tu in ('NONE', '-', 'NAN', ''):
        # Document type WA = jurnal otomatis SAP, teks selalu kosong = normal/SESUAI
        if str(doc).strip().upper() == 'WA':
            return ('SESUAI', '', '')
        return ('TIDAK SESUAI', 'Teks Kosong', '')

    # 2. GL KEPALA 51 & 6 -- cek authorized
    if gl.startswith('51') or gl.startswith('6'):
        if not authorized(username):
            return ('USER TIDAK SAH', 'User Tidak Sah', '')
        return ok()

    # 3. COA tidak di master
    if gl not in coa_map:
        return ok()

    data = coa_map[gl]

    # 4. UT = Uang Transport
    if re.search(r'\bUT\b', tu):
        return ok()

    # ================================================
    #  HANDLER PER COA
    # ================================================

    # -- BBM Niaga (52010101) & BBM Non-Niaga (54010101)
    if gl in ('52010101', '54010101'):
        if 'BBM' in tu or 'SOLAR' in tu or 'BENSIN' in tu:
            return ok()
        if has_nopol(tu):
            if any(k in tu for k in {'POCOAN','POCAON','TAMBAL BAN','KULI',
                                      'GALON','ISI ULANG','SPSI','SBSI'}):
                pass
            else:
                return ok()
        if re.search(r'\d+[,\.]?\d*L\b', tu):
            return ok()

    # -- Parkir & Tol Niaga (52010102)
    if gl == '52010102':
        # [FIX D] KULI/BONGKAR di GL Parkir -> TIDAK SESUAI, saran 52010402
        if any(k in tu for k in {'KULI','BONGKAR','PANGGUL','ANGKUT','BURUH'}):
            return ('TIDAK SESUAI', 'Koreksi COA', coa_label('52010402', coa_map))
        if any(k in tu for k in {'PARKIR','TOL','PKR','PKRNGA','KOORDINATOR',
                                   'SIS','COLL','RETRIBUSI','RPT','BY PARKIR'}):
            return ok()
        if re.search(r'\bPKR\b', tu):
            return ok()

    # -- Retribusi/Keamanan Niaga (52010103)
    if gl == '52010103':
        # [FIX D] KULI/BONGKAR di GL Retribusi -> TIDAK SESUAI, saran 52010402
        if any(k in tu for k in {'KULI','BONGKAR','PANGGUL','ANGKUT','BURUH'}):
            return ('TIDAK SESUAI', 'Koreksi COA', coa_label('52010402', coa_map))
        if any(k in tu for k in {'KMNN','KEAMANAN','MEL','RETRIBUSI','SPSI',
                                   'PUNGLI','PREMAN','PUNGUTAN','KEBERSHN'}):
            return ok()
        if has_nopol(tu):
            return ok()

    # -- Ban Niaga (52010201)
    if gl == '52010201':
        if any(k in tu for k in {'BAN','TAMBAL','TB','TUBLES','TUBELESS',
                                   'ANGIN','GANTI BAN','VELG','KEMPIS',
                                   'FLASHER','BOHLAM','LAMPU','POMPA','KEPALA POMPA','SANYO',
                                   'AKI','SETRUM AKI','ACCUMULATOR','ACCU','BATEREI','BATERAI',
                                   'CUCI','POLES','DETAILING','SERVIS','SPOORING','BALANCING',
                                   'TALI','RANTAI','SPROKET','KAMPAS','OLI',
                                   'FILTER','SPARE PART','SPAREPART','ONDERDIL',
                                   'WIPER','KACA','SELANG','RING','BAUT',
                                   'SERVICE','SERVIS','PERBAIKAN'}):
            return ok()
        if re.search(r'\bTB\b', tu):
            return ok()

    # -- Sewa Kendaraan Niaga (52010301 vs 52010309)
    if gl in ('52010301', '52010309'):
        kw_sewa = {'SEWA','ARMADA','TRUK','PICKUP','MOTOR','GOJEK','GRAB',
                   'BLINDVAN','BV','LC','L300','ENGKEL','PERPANJANGAN','MUTASI',
                   'INSIDENTIL','DRIVER','GEROBAK','BECAK','BENTOR','TRAGA',
                   'CDD','CDE','CD','KEND','KENDARAAN',
                   'DOUBLE','KDR','KIRIM',
                   'PENYEBRANGAN','PENYEBERANGAN','MENYEBRANG','FERRY','FERI',
                   'WB','EREG','GMBOX','WINGBOX','INTRA GDG','OWH','INTRA',
                   'RUTE','JALAN','DARI','MENUJU','SIMPANG','KOTA','PONDOK'}
        has_kend = bool(t & kw_sewa) or 'CDD' in tu or 'CDE' in tu
        if not has_kend:
            fz, _ = fuzzy_hit(t, list(kw_sewa))
            has_kend = fz
        if not has_kend:
            return ('TIDAK SESUAI', 'Koreksi COA', saran(gl, tu, doc, coa_map))
        if gl == '52010301' and doc0 != '4':
            return ('TIDAK SESUAI', 'Koreksi COA', coa_label('52010309', coa_map))
        if gl == '52010309' and doc0 == '4':
            return ('TIDAK SESUAI', 'Koreksi COA', coa_label('52010301', coa_map))
        return ok()

    # -- Upah Harian Lepas (52010401)
    if gl == '52010401':
        if any(k in tu for k in {'POCOAN','POCAON','POCOKAN','UPAH','PHL',
                                   'UHL','HELPER','SUPIR','DRIVER','KENEK',
                                   'PENGGANTI','IMPRES','ALLOWANCE'}):
            return ok()

    # -- Ongkos Kuli (52010402)
    if gl == '52010402':
        if 'KULI' in tu:
            return ok()
        if any(k in t for k in {'ONGKOS','BONGKAR','ANGKUT','PANGGUL','BURUH',
                                  'BNGKR','BNGKRMT','KULI_'}) or \
                re.search(r'BNGKR|BONGKR', tu):
            return ok()
        if re.match(r'^KULI[\s_]', tu):
            return ok()

    # -- Intra Gudang (52010403)
    if gl == '52010403':
        if any(k in tu for k in {'INTRA','INTRAGUDANG','ETOLL',
                                   'INBON','GESER','PINDAH','MUTASI GDG'}):
            return ok()

    # -- Pocoan Helper (52010405)
    if gl == '52010405':
        if any(k in t for k in {'POCOAN','HELPER','ALLOWANCE','DRIVER','POCAON'}):
            return ok()

    # -- Intra OWH (52010406)
    if gl == '52010406':
        if any(k in tu for k in {'INTRA','OWH','GDG'}):
            return ok()

    # -- Transportation (52030107)
    if gl == '52030107':
        if any(k in tu for k in {'TRANSPORTATION','TRANSPORT','TNJ','TNJA',
                                   'KIRIM','PENGIRIMAN','DELIVERY'}):
            return ok()

    # -- Selling Support (53000001)
    if gl == '53000001':
        if any(k in tu for k in {'SELLING SUPPORT','SELLING','SUPPORT',
                                   'SAMPLE','BS MC','ALFAMIDI','B2B','FEE',
                                   'AEON','INDOMARET','ALFAMART','SUPERINDO',
                                   'XTRA','EXTRA','PROMO','PROGRAM','KBN',
                                   'DEPLOY','HADIAH','REWARD','BONUS'}):
            return ok()

    # -- Ban Non-Niaga (54010201)
    if gl == '54010201':
        if any(k in tu for k in {'BAN','TAMBAL','TB','TUBLES','ANGIN','VELG',
                                   'AIR ACCU','ACCU','BATRAI','BATERAI',
                                   'KUNCI MOBIL','LAMPU MOBIL'}):
            return ok()
        if re.search(r'\bTB\b', tu):
            return ok()

    # -- Sewa Kend Non-Niaga (54010301)
    if gl == '54010301':
        if any(k in tu for k in {'SEWA','PERPANJANGAN','MUTASI','KOMPENSASI',
                                   'AVANZA','MOTOR','KENDARAAN'}):
            return ok()

    # -- Perijinan Non-Niaga (54010401)
    if gl == '54010401':
        if has_nopol(tu):
            return ok()
        if any(k in tu for k in {'IZIN','IJIN','PERIJIN','CABUT BERKAS',
                                   'STNK','PAJAK KEND','SIM'}):
            return ok()

    # -- Representatif (55010702)
    if gl == '55010702':
        if any(k in tu for k in {'REPRESENTATIF','REPRESENTATIVE','PERJAMUAN',
                                   'PARCEL','POLSEK','LANTAS','POLANTAS'}):
            return ok()

    # -- ATK & Cetakan (55010101)
    if gl == '55010101':
        atk_kw = {'ATK','KERTAS','TINTA','LEM','SOLASI','ISOLATIF','LAKBAN',
                   'STEROFOAM','BUBBLEWRAP','SELOTIP','KARDUS',
                   'PUSH PIN','MASKER','STEPLES','STREPLES','KLIP','PAPER CLIP',
                   'PENGHAPUS','TIPE-X','CORRECTION','WHITEBOARD','SPIDOL',
                   'LAMINATING','BUKU','KWITANSI','BOLPEN','PULPEN',
                   'STEMPEL','STAPLER','PITA','NCR','PLY','CONT','CF','BANDED',
                   'BOILPOIN','PENA','LABEL','CORRECTION','PENGHAPUS',
                   'AMPLOP','ORDNER','MAP','MIKA','GUNTING','HEKTER','CETAKAN',
                   'TIPE','FOLIO'}
        if t & atk_kw or any(k in tu for k in atk_kw):
            return ok()
        if tu.startswith('ATK') or re.match(r'^ATK[\s_\-]', tu):
            return ok()

    # -- Pos & Dokumen (55010103)
    if gl == '55010103':
        if any(k in tu for k in {'FC','FOTO','KOPI','FOTOCOPY','FOTOKOPI',
                                   'KIRIM','KIRIM DOC','DOKUMEN','PRINT','SCAN',
                                   'MATERAI','POS','JILID','PAKET','EKSPEDISI',
                                   'J&T','JNT','LHPU','BPKK','BOP','REWARD','PPD',
                                   'BPKB','JNE','SICEPAT','WAHANA','TIKI',
                                   'ONGKIR','BIAYA KIRIM','KLAIM MT'}):
            return ok()
        if tu.startswith('POS') or re.match(r'^POS[\s_\-]', tu):
            return ok()
        if re.search(r'\b(JNT|JNE|TIKI|SCP|SICEPAT)\s+[A-Z0-9]{6,}', tu):
            return ok()
        if 'DOC' in t or 'BERKAS' in tu or 'DOKUMEN' in tu:
            return ok()

    # -- Iuran RT/Keamanan (55010601)
    if gl == '55010601':
        if any(k in tu for k in {'IURAN','RT','RW','BABINSA','BABINKAMTIBNAS',
                                   'SATGAS','KEAMANAN','KEBERSIHAN','SAMPAH',
                                   'RETRIBUSI','CLEAN','SATPAM','JAGA','MALAM',
                                   'IPL','LINGKUNGAN','WARGA','KAMPUNG',
                                   'PENGELOLAAN','KOMPLEKS'}):
            return ok()

    # -- Internet/Wifi (55020004)
    if gl == '55020004':
        if any(k in tu for k in {'INTERNET','WIFI','INDIHOME','BIZNET','BISNET',
                                   'SIGNA','ICONNET','GASSNET','DGNET','ALFANET',
                                   'KUOTA','PULSA','FIRSTMEDIA','VIGO',
                                   'JDN','ORBIT','DATA','MODEM','PAKET DATA',
                                   'PROVIDER','ZKM','JARINGAN'}):
            return ok()

    # -- Telepon & Pulsa (55020003)
    if gl == '55020003':
        if any(k in tu for k in {'PULSA','TELEPON','TELPON','TELFON','PHONE',
                                   'KUOTA','INTERNET','WIFI','KANTOR','HANDPHONE',
                                   'HP','BY TELEPON'}):
            return ok()
        if re.search(r'PULSA\s+\d{8,}', tu) or 'PULSA KANTOR' in tu:
            return ok()

    # -- Rumah Tangga Kantor (55010201)
    if gl == '55010201':
        if any(k in tu for k in {'TRANSP KSR','TRANSP KASIR','TRANSPORT KASIR'}):
            return ok()
        if re.search(r'\bKASIR\b', tu) or re.search(r'\bSETOR\b', tu):
            return ok()
        if any(k in tu for k in {'GRAB','GOJEK','OJEK'}) and \
           any(k in tu for k in {'BANK','SETOR','SETORAN','BRILINK','BRI',
                                   'MANDIRI','BCA','BNI','BSI'}):
            return ok()
        if 'GENSET' in tu:
            return ok()
        if any(k in tu for k in {'MUSNAH BS','BAKAR BS','BUANG BS','BI MUSNAH',
                                   'PEMBAKARAN','PEMUSNAHAN'}):
            return ok()

    # -- [FIX A] Security (55050001): hanya P000 & SAP_SYSTEM
    if gl == '55050001':
        m_u    = re.match(r'^P(\d+)', str(username).upper())
        kode_u = int(m_u.group(1)) if m_u else 9999
        is_auth = (kode_u == 0) or str(username).upper().startswith('SAP_')
        if not is_auth:
            return ('TIDAK SESUAI', 'Koreksi COA',
                    coa_label('55050009', coa_map))
        return ok()

    # -- Pemeliharaan Bangunan (55030001)
    if gl == '55030001':
        kw_bang = {'LAMPU','KABEL','BANGUNAN','TOILET','SALURAN','TUKANG',
                   'BOHLAM','PERBAIKAN','WATER PUMP','BUILDING MAINTENANCE',
                   'GRENDEL','PIPA','PARALON','KERAN','MESIN AIR','POMPA AIR',
                   'WATER','PUMP','MAINTENANCE','CAT','TEMBOK','PLAFON',
                   'LANTAI','GENTENG','PINTU','JENDELA','KUNCI','GEMBOK',
                   'BESI','SEMEN','PASIR','RENOVASI','POMPA','SANYO','OTOMATIS',
                   'INSTALASI','LISTRIK','SAKLAR','STOP KONTAK','CABLE'}
        if t & kw_bang or any(k in tu for k in kw_bang):
            return ok()

    # -- Pemeliharaan Alat Kantor (55030002)
    if gl == '55030002':
        kw_alat = {'SERVICE','SERVIS','PERBAIKAN','CUCI','GENSET','AC','PRINTER',
                   'MESIN','FINGER','BOCOR','KAPASITOR','SRVC','MOUSE','KEYBOARD',
                   'MONITOR','KOMPUTER','LAPTOP','UPS','DISPENSER','POMPA','HEAD',
                   'MAINTENANCE','EQUIPMENT','OFFICE EQUIPMENT','PERAWATAN',
                   'WATER PUMP','MESIN AIR','POMPA AIR'}
        if t & kw_alat or any(k in tu for k in kw_alat):
            return ok()

    # -- OB / Cleaning (55050009)
    if gl == '55050009':
        if any(k in tu for k in {'OB','WAKAR','ART','OUTSOURCING','CLEANING',
                                   'KEBERSIHAN','GAJI OB','OFFICE BOY',
                                   'CENTENG','OG','OGAH'}):
            return ok()
        if re.search(r'\bGAJI\s+O[BG]\b', tu) or re.search(r'\bO[BG]\s+\d', tu):
            return ok()

    # -- Hotel/Penginapan PPD (55080201)
    if gl == '55080201':
        if any(k in tu for k in {'KOS','KOST','HOTEL','HTL','PENGINAPAN',
                                   'INAP','MESS','WISMA','LPPD','PPD','AKO',
                                   'BY HOTEL','HTLSLS','KONTRAKAN','SEWA RUMAH',
                                   'KOST-KOSTAN','PONDOKAN','LOSMEN'}):
            return ok()

    # -- BBM Perjalanan Dinas (55080102)
    if gl == '55080102':
        if re.match(r'^BO[\s_\-]', tu) or re.search(r'BO', tu):
            return ok()
        has_bbm = 'BBM' in tu or 'SOLAR' in tu or has_nopol(tu)
        has_ppd = any(k in tu for k in {'PPD','PERDIN','PERJALANAN DINAS',
                                          'ADC PPD','BBMPERDIN','ADC'})
        if has_bbm and has_ppd:
            return ok()
        if 'BBM PERJALANAN DINAS' in dhtu or 'BBM PERDIN' in dhtu:
            return ok()
        if has_bbm and not has_ppd:
            return ('TIDAK SESUAI', 'Koreksi COA', coa_label('52010101', coa_map))

    # -- Tiket (55080101)
    if gl == '55080101':
        if any(k in tu for k in {'TIKET','TKT','TICKET','PESAWAT','KERETA',
                                   'KAPAL','TRAVEL','PENYEBRANGAN','PENYEBERANGAN',
                                   'FERRY','FERI','SPITT','BY.SPITT'}):
            return ok()
        if re.match(r'^BY\.?\s+\w+', tu):
            return ok()

    # -- Tol/Parkir PPD (55080103)
    if gl == '55080103':
        if any(k in tu for k in {'TOL','PARKIR','PKR','ADC PPD TOL',
                                   'TOLL PPD','LPPD','BO'}):
            return ok()

    # -- Uang Makan Perdin (55080203)
    if gl == '55080203':
        if any(k in tu for k in {'UM','UANG MKN','UANG MAKAN','UMDIN',
                                   'PPD','PERDIN','MAKAN','MAKAN SIANG'}):
            return ok()

    # -- Transport/Sewa PPD (55080301)
    if gl == '55080301':
        if any(k in tu for k in {'PPD','PERDIN','SEWA MOBIL','TAXI','TAKSI',
                                   'GOJEK','GRAB','SEWA MOTOR','SW MTR',
                                   'TRANSPORT','TRANSPORTASI','PPD_LL',
                                   'PERJALANAN DINAS','PERBANTUAN','PENDAMPINGAN',
                                   'DRIVER','SUPIR'}):
            return ok()
        if has_nopol(tu) and any(k in tu for k in {'PPD','PERDIN','ADC'}):
            return ok()

    # -- Biaya Scrapping (57000001)
    if gl == '57000001':
        if any(k in tu for k in {'SCRAP','SCRAPPING','SCREAPPING','PENYUSUTAN',
                                   'WRITE OFF','PENGHAPUSAN','DISPOSAL'}):
            return ok()
        if re.match(r'^P\d{3}/', tu):
            return ok()
        if re.match(r'^(JAN|FEB|MAR|APR|MEI|JUN|JUL|AGU|SEP|OKT|NOV|DES)\s+20\d{2}', tu):
            return ('SESUAI', 'COA-Text-Header Tidak Relevan', '')

    # -- Pendapatan Bunga (71010001)
    if gl == '71010001':
        if any(k in tu for k in {'ADM','SWEEP','UBP','ABODEMEN','VIRTUAL ACC',
                                   'KLIRING','CHARGE','VIRTUAL','JASA GIRO',
                                   'BUNGA','DEPOSITO','BAGI HASIL'}):
            return ok()

    # -- Denda/Penalty (71040001)
    if gl == '71040001':
        if any(k in tu for k in {'PENALTY','PENALTI','DENDA','DN','SANKSI'}):
            return ok()
        if 'PENALTY' in dhtu or 'PENALTI' in dhtu or 'DENDA' in dhtu:
            return ok()
        if re.match(r'^DN\s+\d+/', tu):
            return ok()

    # -- Biaya Adm Bank (72010001)
    if gl == '72010001':
        kw72 = {'ADMIN','ADM','TRANSFER','TF','TFR','MATERAI','BY','BIAYA',
                'KLIRING','SETOR','BRILINK','BRI','ABODEMEN','CHARGE','SWEEP',
                'VIRTUAL ACC','VIRTUAL','ATM','UBP','ADMINISTRASI','JASA GIRO',
                'BUNGA','ABONEMEN','BANK','TRF','SETORAN','VA','KORAN',
                'ADJUST','SELISIH','KOREKSI','REKENING','NOTA DEBET','ND'}
        if t & kw72 or any(k in tu for k in kw72):
            return ok()
        if re.search(r'\b(TF|TFR|ADM|BY)\b', tu):
            return ok()

    # -- 71060001 vs 71060002 (berdasarkan amount)
    if gl in ('71060001', '71060002'):
        if gl == '71060001':
            if absa > 100:
                return ('TIDAK SESUAI', 'Koreksi COA',
                        '71060002 - Pendapatan/Kerugian Lainnya (nominal > Rp100)')
            return ok()
        else:
            if absa <= 100:
                return ('TIDAK SESUAI', 'Koreksi COA',
                        '71060001 - Selisih Lebih/Kurang (<=Rp100)')
            return ok()

    # -- Fuzzy match terhadap master COA token
    common = t & data['tokens']
    if common:
        return ok()

    fz2, _ = fuzzy_hit(t, list(data['tokens']))
    if fz2:
        return ok()

    # TIDAK SESUAI
    sr = saran(gl, tu, doc, coa_map)
    return ('TIDAK SESUAI', 'Koreksi COA', sr)

# ==============================
#  KLASIFIKASI NON-BEBAN
# ==============================
def klas_ap(gl):
    return {'1':'AKTIVA','2':'PASIVA','3':'EKUITAS','4':'PENDAPATAN'}.get(gl[0],'NON-BEBAN')

# ==============================
#  PROSES SATU SHEET
# ==============================
def proses_sheet(ws_src, ws_dst, coa_map, user_map, fsz, row_h, tab_color=None,
                  freeze=None, cc_index=None):
    cc_index = cc_index if cc_index is not None else build_cc_index(user_map)
    if tab_color:
        ws_dst.sheet_properties.tabColor = OXColor(rgb=tab_color)
    if freeze:
        ws_dst.freeze_panes = freeze

    orig_hdrs = [c.value for c in ws_src[1]]
    n_col     = len(orig_hdrs)

    # Header
    for ci, val in enumerate(['STATUS','REMARKS','SARAN COA','KESESUAIAN COST CENTER','CHANNEL'] + orig_hdrs, 1):
        c = ws_dst.cell(row=1, column=ci, value=val)
        c.font = fnt(fsz); c.fill = FILL_HDR
        c.alignment = ALIGN_TOP; c.border = BORDER_HDR
    ws_dst.row_dimensions[1].height = row_h

    # Lebar kolom
    for cl, dim in ws_src.column_dimensions.items():
        from openpyxl.utils import column_index_from_string
        ci = column_index_from_string(cl)
        ws_dst.column_dimensions[get_column_letter(ci+5)].width = dim.width
    ws_dst.column_dimensions['D'].width = 26
    ws_dst.column_dimensions['A'].width = 15
    ws_dst.column_dimensions['B'].width = 28
    ws_dst.column_dimensions['C'].width = 35
    ws_dst.column_dimensions['E'].width = 12

    stats = {'SESUAI':0,'TIDAK SESUAI':0,'USER TIDAK SAH':0,'NON-BEBAN':0,
              'SUDAH_DIKOREKSI':0,'BELUM_DIKOREKSI':0}
    daftar_koreksi = []
    rows  = list(ws_src.iter_rows(min_row=2, values_only=True))

    # -- Pre-pass: index semua dokumen (dipakai utk mencari pasangan reversal
    #    di bawah -- lihat "Reversal Clear"). Murni pembacaan data, tidak
    #    menulis apa pun ke Excel, tidak memengaruhi evaluate()/validate_dht().
    dok_index = {}
    for row in rows:
        doc_i = str(row[4]).strip() if row[4] else ''
        if not doc_i:
            continue
        dok_index[doc_i] = {
            'gl'      : str(int(row[0])) if isinstance(row[0],(int,float)) else str(row[0]).strip() if row[0] else '',
            'cc'      : str(row[16]).strip() if len(row) > 16 and row[16] and str(row[16]) not in ('None','') else '',
            'txtp'    : str(row[15]).strip() if len(row) > 15 and row[15] else '',
            'dhtp'    : str(row[18]).strip() if len(row) > 18 and row[18] and str(row[18]) not in ('None','') else '',
            'amt'     : row[11] if isinstance(row[11], (int, float)) else 0,
            'rev_dgn' : str(row[21]).strip() if len(row) > 21 and row[21] and str(row[21]) not in ('None','') else '',
        }

    for ri, row in enumerate(rows, start=2):
        gl   = str(int(row[0])) if isinstance(row[0],(int,float)) else str(row[0]).strip() if row[0] else ''
        user = str(row[1]).strip() if row[1] else ''
        doc      = str(row[4]).strip() if row[4] else ''
        doc_type = str(row[6]).strip() if len(row) > 6 and row[6] else ''
        amt  = row[11] if isinstance(row[11],(int,float)) else 0
        txtp = str(row[15]).strip() if len(row)>15 and row[15] else ''

        dhtp     = str(row[18]).strip() if len(row) > 18 and row[18] and str(row[18]) not in ('None','') else ''
        dhtu     = dhtp.upper()
        cc_fbl3n = str(row[16]).strip() if len(row) > 16 and row[16] and str(row[16]) not in ('None','') else ''
        # Kolom V "Reversed With" -- terisi begitu dokumen ini SUDAH DIBALIK
        # (reversal) di SAP. Dashboard Biaya memakai ini sbg penanda otomatis
        # "sudah dikoreksi" utk baris TIDAK SESUAI/USER TIDAK SAH -- staff
        # SELALU mengoreksi lewat reversal + posting baru, tidak pernah
        # mengedit dokumen lama di tempat.
        rev_dgn  = str(row[21]).strip() if len(row) > 21 and row[21] and str(row[21]) not in ('None','') else ''

        if gl and gl[0] >= '5':
            status, rm, sr = evaluate(gl, txtp, doc_type, user, amt, coa_map, dhtp)
            stats[status] = stats.get(status, 0) + 1
        elif gl and gl[0] in '1234':
            stats['NON-BEBAN'] += 1
            continue
        else:
            continue

        # Remarks logic
        saran_val = '' if (not sr or sr in ('-', '')) else sr
        if rm == 'Koreksi COA' and not saran_val:
            rm = 'Cek Standar Penulisan'

        # Warna
        fill = (FILL_NO    if status == 'TIDAK SESUAI'              else
                FILL_AKSES if status == 'USER TIDAK SAH'            else
                FILL_DHT   if rm == 'COA-Text-Header Tidak Relevan' else
                FILL_OK)

        # Kolom A = STATUS
        ca = ws_dst.cell(row=ri, column=1, value=status)
        ca.font = fnt(fsz); ca.fill = fill
        ca.alignment = ALIGN_TOP; ca.border = BORDER_NON

        # Kolom B = REMARKS
        rm_fill = (FILL_DHT  if rm == 'COA-Text-Header Tidak Relevan' else
                   FILL_WARN  if rm == 'Teks Kosong'                   else
                   FILL_FMT   if rm == 'Cek Standar Penulisan'         else
                   FILL_NONE)
        cb = ws_dst.cell(row=ri, column=2, value=rm)
        cb.font = fnt(fsz); cb.fill = rm_fill
        cb.alignment = ALIGN_TOP; cb.border = BORDER_NON

        # Kolom C = SARAN COA
        saran_out = saran_val if status == 'TIDAK SESUAI' else ''
        cs = ws_dst.cell(row=ri, column=3, value=saran_out)
        cs.font = fnt(fsz); cs.fill = FILL_NONE
        cs.alignment = ALIGN_TOP; cs.border = BORDER_NON

        # Kolom D = KESESUAIAN COST CENTER
        cc_val, cc_fill, channel_val, paksa_status = cek_cc(user, cc_fbl3n, user_map, cc_index)

        if paksa_status == 'USER TIDAK SAH' and status != 'USER TIDAK SAH':
            # Profit Center-nya ADA di master tapi statusnya Non Aktif --
            # ini pelanggaran lebih kuat daripada sekadar CC tidak sesuai,
            # jadi menang/override apa pun status hasil evaluate() di atas
            # (termasuk kalau sebelumnya sudah TIDAK SESUAI karena alasan lain).
            stats[status] = stats.get(status, 0) - 1
            status = 'USER TIDAK SAH'
            fill   = FILL_AKSES
            ca.value = status
            ca.fill  = fill
            stats['USER TIDAK SAH'] = stats.get('USER TIDAK SAH', 0) + 1
        # Jika CC tidak sesuai → STATUS kolom A ikut jadi TIDAK SESUAI
        # (kecuali sudah USER TIDAK SAH, biarkan)
        elif cc_fill == FILL_CC_NO and status == 'SESUAI':
            status = 'TIDAK SESUAI'
            fill   = FILL_NO
            # Update kolom A yang sudah ditulis
            ca.value = status
            ca.fill  = fill
            # Ikut turunkan hitungan ringkasan (stats) supaya konsol &
            # sheet SUMMARY tetap konsisten dengan status final di kolom A
            # -- sebelumnya stats sudah kadung dihitung SESUAI di atas
            # sebelum baris ini tahu Cost Center-nya tidak sesuai.
            stats['SESUAI'] -= 1
            stats['TIDAK SESUAI'] = stats.get('TIDAK SESUAI', 0) + 1

        # -- Reversal Clear ---------------------------------------------
        # Kalau baris ini TIDAK SESUAI (bukan USER TIDAK SAH -- itu soal
        # otorisasi, bukan soal isi transaksi, jadi tidak ikut di-clear di
        # sini) DAN sudah punya pasangan reversal yang: (1) saling menunjuk
        # balik satu sama lain, (2) GL/Cost Center/Text/Document Header Text
        # SAMA persis, (3) nominal plus-minus = 0 -- dianggap CLEAR, sudah
        # tidak ada jurnal netto yang berdiri, jadi dikembalikan jadi SESUAI.
        # User Name SENGAJA tidak dicek sama (kasir vs supervisor yg
        # reversal biasa beda orang, itu wajar).
        # status_asal disimpan SEBELUM override ini -- dipakai supaya baris
        # yang di-clear TETAP kelihatan di monitoring SUMMARY (Cuma status
        # akhir di kolom A yang SESUAI, bukan berarti "tidak pernah ada
        # masalah" -- itu beda hal).
        status_asal = status
        remarks_asal = rm
        if status == 'TIDAK SESUAI' and rev_dgn:
            partner = dok_index.get(rev_dgn)
            if (partner and partner['rev_dgn'] == doc
                    and partner['gl'] == gl and partner['cc'] == cc_fbl3n
                    and partner['txtp'] == txtp and partner['dhtp'] == dhtp
                    and abs((amt or 0) + (partner['amt'] or 0)) < 0.01):
                stats['TIDAK SESUAI'] -= 1
                status = 'SESUAI'
                fill   = FILL_OK
                rm     = ''
                saran_out = ''
                ca.value = status; ca.fill = fill
                cb.value = rm;     cb.fill = FILL_NONE
                cs.value = saran_out
                stats['SESUAI'] = stats.get('SESUAI', 0) + 1

        cd = ws_dst.cell(row=ri, column=4, value=cc_val)
        cd.font = fnt(fsz)
        cd.fill = cc_fill if cc_fill else FILL_NONE
        cd.alignment = ALIGN_TOP; cd.border = BORDER_NON

        # Kolom E = CHANNEL
        ce = ws_dst.cell(row=ri, column=5, value=channel_val)
        ce.font = fnt(fsz); ce.fill = FILL_NONE
        ce.alignment = ALIGN_TOP; ce.border = BORDER_NON

        # Kolom F dst = data asli
        for ci, val in enumerate(row, start=6):
            co = ws_dst.cell(row=ri, column=ci, value=val)
            co.font = fnt(fsz); co.fill = FILL_NONE
            co.alignment = ALIGN_TOP; co.border = BORDER_NON
        ws_dst.row_dimensions[ri].height = row_h

        # -- Monitoring koreksi (reversal) -----------------------------------
        # Dicek pakai status_asal (SEBELUM Reversal Clear) supaya baris yang
        # sudah di-clear balik jadi SESUAI TETAP kelihatan di sini (sbg kasus
        # yang sudah selesai), bukan menghilang seolah tidak pernah ada
        # masalah. Kolom STATUS di FBL3N/FBL3N ZBA sendiri tidak terpengaruh
        # -- ini murni catatan tambahan di sheet SUMMARY.
        if status_asal in ('TIDAK SESUAI', 'USER TIDAK SAH'):
            auto_cleared = (status == 'SESUAI')
            sudah = bool(rev_dgn)
            stats['SUDAH_DIKOREKSI' if sudah else 'BELUM_DIKOREKSI'] = \
                stats.get('SUDAH_DIKOREKSI' if sudah else 'BELUM_DIKOREKSI', 0) + 1
            daftar_koreksi.append({
                'sheet': ws_dst.title,
                'baris_sumber': ri, 'nomor_dokumen': doc, 'gl_account': gl,
                'user': user, 'nominal': amt, 'status': status_asal, 'remarks': remarks_asal,
                'sudah_dikoreksi': sudah, 'direversal_dengan': rev_dgn,
                'auto_cleared': auto_cleared, 'status_akhir': status,
            })

        if ri % 2000 == 0:
            print(f'      ... {ri-1:,} baris')

    stats['TOTAL'] = len(rows)
    return stats, daftar_koreksi

# ==============================
#  BUAT SUMMARY
# ==============================
def buat_summary(wb_out, s1, s2, daftar_koreksi):
    ws = wb_out.create_sheet('SUMMARY')
    ws.sheet_properties.tabColor = OXColor(rgb='FFBDD6EE')

    def tabel(sc, judul, stats):
        tot = stats['TOTAL']
        c = ws.cell(row=1, column=sc, value=judul)
        c.font = fnt(); c.fill = FILL_HDR; c.alignment = ALIGN_TOP; c.border = BORDER_HDR
        ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=sc+2)
        for i, h in enumerate(['STATUS','JUMLAH','% TOTAL']):
            c2 = ws.cell(row=2, column=sc+i, value=h)
            c2.font = fnt(); c2.fill = FILL_HDR; c2.alignment = ALIGN_TOP
            c2.border = BORDER_HDR if i == 0 else BORDER_NON
        rows = [
            ('SESUAI',         stats.get('SESUAI',0),         FILL_OK),
            ('TIDAK SESUAI',   stats.get('TIDAK SESUAI',0),   FILL_NO),
            ('USER TIDAK SAH', stats.get('USER TIDAK SAH',0), FILL_AKSES),
            ('NON-BEBAN',      stats.get('NON-BEBAN',0),      FILL_NONE),
            ('TOTAL',          tot,                            PatternFill('solid', fgColor='C0C0C0')),
        ]
        for ri, (lab, jml, fl) in enumerate(rows, start=3):
            pct = f'{jml/tot*100:.1f}%' if tot else '0.0%'
            for ci, val in enumerate([lab, jml, pct]):
                c3 = ws.cell(row=ri, column=sc+ci, value=val)
                c3.font = fnt(); c3.fill = fl; c3.alignment = ALIGN_TOP
                c3.border = BORDER_HDR if ci == 0 else BORDER_NON
        for r in range(1, 10):
            ws.row_dimensions[r].height = ROW_H1
        ws.column_dimensions[get_column_letter(sc)].width   = 15
        ws.column_dimensions[get_column_letter(sc+1)].width = 10
        ws.column_dimensions[get_column_letter(sc+2)].width = 9

    tabel(1, 'FBL3N',     s1)
    ws.column_dimensions['D'].width = 3
    tabel(5, 'FBL3N ZBA', s2)

    # ==========================================================
    #  MONITORING KOREKSI (REVERSAL) -- prinsip sama seperti
    #  halaman "Transaksi Tidak Sesuai" di Dashboard Biaya dulu:
    #  kolom SAP "Reversed With" dipakai sbg penanda otomatis
    #  "sudah dikoreksi", murni informasi tambahan, tidak mengubah
    #  status/warna apa pun di sheet FBL3N/FBL3N ZBA.
    #
    #  Dihitung per KASUS, bukan per baris -- 1 pasangan reversal (2 baris
    #  yang saling menunjuk) dihitung SEKALI sbg "Sudah Dikoreksi", supaya
    #  tidak dobel (mis. 4 baris = 2 pasangan -> tampil 2, bukan 4).
    # ==========================================================
    r0 = 10
    kasus = {}  # key -> {'sudah': bool, 'items': [...]}
    for k in daftar_koreksi:
        if k['direversal_dengan']:
            key = ('pair', k['sheet'], tuple(sorted([k['nomor_dokumen'], k['direversal_dengan']])))
        else:
            key = ('single', k['sheet'], k['baris_sumber'])
        if key not in kasus:
            kasus[key] = {'sudah': bool(k['direversal_dengan']), 'items': []}
        kasus[key]['items'].append(k)

    tot_bermasalah = len(kasus)
    sudah = sum(1 for v in kasus.values() if v['sudah'])
    belum = tot_bermasalah - sudah

    c = ws.cell(row=r0, column=1,
                value='MONITORING KOREKSI (REVERSAL) - Baris TIDAK SESUAI / USER TIDAK SAH')
    c.font = fnt(); c.fill = FILL_HDR; c.alignment = ALIGN_TOP; c.border = BORDER_HDR
    ws.merge_cells(start_row=r0, start_column=1, end_row=r0, end_column=3)
    ws.row_dimensions[r0].height = ROW_H1

    ringkas = [
        ('Sudah Dikoreksi (Reversal)', sudah, FILL_OK),
        ('Belum Dikoreksi (Gantungan)', belum, FILL_NO),
        ('Total Kasus Bermasalah', tot_bermasalah, PatternFill('solid', fgColor='C0C0C0')),
    ]
    for i, (lab, jml, fl) in enumerate(ringkas, start=1):
        pct = f'{jml/tot_bermasalah*100:.1f}%' if tot_bermasalah else '0.0%'
        rr = r0 + i
        for ci, val in enumerate([lab, jml, pct]):
            c3 = ws.cell(row=rr, column=1+ci, value=val)
            c3.font = fnt(); c3.fill = fl; c3.alignment = ALIGN_TOP
            c3.border = BORDER_HDR if ci == 0 else BORDER_NON
        ws.row_dimensions[rr].height = ROW_H1

    # -- Daftar detail: masih per BARIS (bukan per kasus) supaya tiap baris
    #    tetap bisa ditelusuri satu-satu -- gantungan didahulukan supaya
    #    langsung kelihatan apa yang masih perlu ditindaklanjuti. Baris yang
    #    sudah AUTO-SESUAI (lolos aturan Reversal Clear) TETAP muncul di
    #    sini (statusnya "asal", bukan status akhir di kolom A) supaya
    #    kelihatan riwayatnya, bukan menghilang begitu saja.
    daftar_urut = sorted(daftar_koreksi, key=lambda k: (k['sudah_dikoreksi'], k['sheet'], k['baris_sumber']))
    rh = r0 + len(ringkas) + 2
    judul = ws.cell(row=rh, column=1, value='DAFTAR DETAIL (gantungan ditampilkan lebih dulu)')
    judul.font = fnt(); judul.fill = FILL_HDR; judul.alignment = ALIGN_TOP; judul.border = BORDER_HDR
    ws.merge_cells(start_row=rh, start_column=1, end_row=rh, end_column=10)
    ws.row_dimensions[rh].height = ROW_H1

    headers = ['SHEET','BARIS SUMBER','NOMOR DOKUMEN','GL ACCOUNT','USER',
               'NOMINAL','STATUS ASAL','REMARKS ASAL','STATUS AKHIR','STATUS KOREKSI']
    hr = rh + 1
    for ci, h in enumerate(headers, start=1):
        c2 = ws.cell(row=hr, column=ci, value=h)
        c2.font = fnt(); c2.fill = FILL_HDR; c2.alignment = ALIGN_TOP; c2.border = BORDER_HDR
    ws.row_dimensions[hr].height = ROW_H1

    for i, k in enumerate(daftar_urut, start=1):
        rr = hr + i
        if k['auto_cleared']:
            status_koreksi = f"Sudah Dikoreksi -> Otomatis SESUAI (Reversal: {k['direversal_dengan']})"
        elif k['sudah_dikoreksi']:
            status_koreksi = f"Sudah Dikoreksi, tapi field tidak identik -- cek manual (Reversal: {k['direversal_dengan']})"
        else:
            status_koreksi = 'Belum Dikoreksi (Gantungan)'
        vals = [k['sheet'], k['baris_sumber'], k['nomor_dokumen'], k['gl_account'],
                 k['user'], k['nominal'], k['status'], k['remarks'], k['status_akhir'], status_koreksi]
        fl = FILL_OK if k['sudah_dikoreksi'] else FILL_NO
        for ci, val in enumerate(vals, start=1):
            c3 = ws.cell(row=rr, column=ci, value=val)
            c3.font = fnt(); c3.fill = fl; c3.alignment = ALIGN_TOP; c3.border = BORDER_NON
        ws.row_dimensions[rr].height = ROW_H1

    for ci, w in enumerate([12, 12, 16, 11, 14, 14, 14, 22, 14, 42], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = f'A{hr+1}'

# ==============================
#  MAIN
# ==============================
def proses():
    print('='*60)
    print('  VALIDASI COA FBL3N vs MASTER COA  [v6 - LOKAL]')
    print('='*60)
    print(f'\n  Input  : {INPUT_FBL3N}')
    print(f'  Master : {INPUT_MASTER}')
    print(f'  User   : {INPUT_USER if os.path.exists(INPUT_USER) else INPUT_USER + " [TIDAK DITEMUKAN]"}')
    print(f'  Output : {OUTPUT_FILE}\n')

    for path, label in [(INPUT_FBL3N,'FBL3N.xlsx'),(INPUT_MASTER,'MASTER_COA.xlsx')]:
        if not os.path.exists(path):
            print(f'[ERROR] File tidak ditemukan: {path}')
            input('Tekan Enter untuk keluar...')
            return

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print('[1/5] Memuat MASTER COA & USER...')
    coa_map  = load_master(INPUT_MASTER)
    user_map = load_master_user(INPUT_USER)
    print(f'      {len(coa_map)} COA termuat, {len(user_map)} user termuat\n')

    print('[2/5] Memuat FBL3N.xlsx...')
    with open(INPUT_FBL3N, 'rb') as f:
        raw = f.read()
    wb_src = openpyxl.load_workbook(io.BytesIO(raw))
    print(f'      Sheet: {wb_src.sheetnames}\n')

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    print('[3/5] Memproses sheet FBL3N...')
    ws1 = wb_out.create_sheet('FBL3N')
    s1, k1 = proses_sheet(wb_src['FBL3N'], ws1, coa_map, user_map,
                       fsz=10, row_h=ROW_H1, tab_color='FFBDD6EE', freeze='B2')
    print(f'      {s1["TOTAL"]:,} baris | '
          f'SESUAI={s1["SESUAI"]:,} | '
          f'TIDAK SESUAI={s1["TIDAK SESUAI"]:,} | '
          f'USER TDK SAH={s1["USER TIDAK SAH"]:,}\n')

    print('[4/5] Memproses sheet FBL3N ZBA...')
    ws2 = wb_out.create_sheet('FBL3N ZBA')
    s2, k2 = proses_sheet(wb_src['FBL3N ZBA'], ws2, coa_map, user_map,
                       fsz=11, row_h=ROW_H2, tab_color=None, freeze=None)
    print(f'      {s2["TOTAL"]:,} baris | NON-BEBAN={s2["NON-BEBAN"]:,}\n')

    print('[5/5] Membuat SUMMARY & menyimpan...')
    buat_summary(wb_out, s1, s2, k1 + k2)
    wb_out.save(OUTPUT_FILE)
    wb_src.close()

    print()
    print('='*60)
    print('  SELESAI')
    print('='*60)
    print(f'  File     : {OUTPUT_FILE}')
    print(f'  Sheet    : FBL3N | FBL3N ZBA | SUMMARY')
    print()
    for label, stats in [('FBL3N', s1), ('FBL3N ZBA', s2)]:
        print(f'  {label} ({stats["TOTAL"]:,} baris):')
        print(f'    SESUAI          : {stats.get("SESUAI",0):,}')
        print(f'    TIDAK SESUAI    : {stats.get("TIDAK SESUAI",0):,}')
        print(f'    USER TIDAK SAH  : {stats.get("USER TIDAK SAH",0):,}')
        print(f'    NON-BEBAN       : {stats.get("NON-BEBAN",0):,}')
        print()
    print('='*60)

    # -- Salin ke folder kerja divisi di drive D: ----------------------------
    # OUTPUT .xlsx di atas SUDAH tersimpan sebelum baris ini dijalankan --
    # gagal/dilewatinya langkah ini (mis. drive D: tidak ada) tidak pernah
    # menghapus/mengubah hasil validasi yang sudah ada di disk lokal.
    # Ini menggantikan sinkronisasi ke Supabase/website versi sebelumnya --
    # tim lain (mis. P&L) tinggal buka folder ini langsung, tanpa login web.
    try:
        tujuan_root = os.path.join(ROOT_PROJECT_FAD, DIVISI_TUJUAN, FOLDER_TUJUAN)
        if os.path.abspath(tujuan_root) == os.path.abspath(BASE_DIR):
            # Script ini sudah dijalankan langsung di dalam folder tujuan
            # (mis. tool ditaruh langsung di folder kerja divisi di drive D:).
            # Output lokal di atas SUDAH berada di lokasi yang benar -- tidak
            # perlu disalin lagi ke tempat yang sama (menghindari folder
            # duplikat di dalam OUTPUT\).
            pass
        elif os.path.exists(ROOT_PROJECT_FAD.split(':')[0] + ':\\'):
            # Setiap run dapat folder sendiri di dalam OUTPUT\ (dinamai
            # sesuai waktu proses) -- supaya file .xlsx tidak menumpuk jadi
            # satu tumpukan berantakan di folder kerja divisi.
            tujuan_dir = os.path.join(tujuan_root, "OUTPUT", _TS)
            os.makedirs(tujuan_dir, exist_ok=True)
            tujuan_file = os.path.join(tujuan_dir, os.path.basename(OUTPUT_FILE))
            shutil.copy2(OUTPUT_FILE, tujuan_file)
            print(f'  [OK] Salinan output disimpan juga di:')
            print(f'       {tujuan_file}')
        else:
            print(f'  [PERINGATAN] Drive D: tidak ditemukan -- salinan ke folder')
            print(f'       divisi dilewati. Output lokal di OUTPUT\\ tetap aman.')
    except Exception as exc:
        print(f'  [PERINGATAN] Gagal menyalin ke folder divisi ({exc}).')
        print(f'       Output lokal di OUTPUT\\ tetap aman, tidak terpengaruh.')

    input('\n  Tekan Enter untuk keluar...')

if __name__ == '__main__':
    proses()
